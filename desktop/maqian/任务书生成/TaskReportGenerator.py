# coding=utf-8

import shutil

from openpyxl import load_workbook


wb = load_workbook(ur'任务书模板.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
rows = list(sheet.rows)
heads = rows[0]
heads = map(lambda cell: cell.value, heads)

for row in rows[1:]:
    cells = map(lambda c: c.value, row)
    destFileName = u'{0}-{1}'.format(cells[1], cells[5])
    destFilePath = ur'./taskReport/{0}.xlsx'.format(destFileName)
    shutil.copyfile(u'任务书-通用.xlsx', destFilePath)

    newWb = load_workbook(destFilePath)
    newSheet = newWb.get_sheet_by_name('Sheet1')
    locationValDict = {'E4': 'G', 'J4': 'F', 'C5': 'H', 'J5': 'I', 'H7': 'J', 'O7': 'K', 'C8': 'L', 'O9': 'M', 'E10': 'C',
                       'O10': 'D', 'H11': 'E', 'O20': 'N', 'P20': 'O', 'Q20': 'P', 'O21': 'Q', 'O22': 'R', 'O24': 'S'}
    for location, valueIndex in locationValDict.items():
        value = cells[ord(valueIndex) - 65]
        newSheet[location] = value
    newSheet['B9'] = cells[0].year
    newSheet['D9'] = cells[0].month
    newSheet['F9'] = cells[0].day
    newWb.save(destFilePath)
