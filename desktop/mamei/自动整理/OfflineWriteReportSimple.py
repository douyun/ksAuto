# coding=utf-8
import csv
import os
import re
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

SrcDestKeyDict = {u'检验日期': u'检验日期',
                  u'服务地点': u'服务地点',
                  u'零件名称': u'零件名称',
                  u'零件号': u'零件号',
                  u'服务明细': u'服务明细',
                   u'批次号': u'批次号',
                   u'供应商批次': u'供应商批次',
                   u'检验总数': u'检验总数',
                   u'任务人员': u'任务人员',
                   u'任务时间': u'任务时间',
                   u'效率': u'效率',
                  u'任务工时': u'任务工时',
                  u'班次': u'班次',
                  u'不良数量': u'不良数量',
                  u'不良类型': u'不良明细'}

MergeCols = [u'任务工时', u'任务时间', u'任务人员', u'效率']

FormulaDict = {u'合格数量': u'检验总数-不良数量'}

OverallFormulaDict = {u'检验总数': '=SUM({0}{1}:{0}{2})',
                      u'不良数量': '=SUM({0}{1}:{0}{2})',
                      u'任务工时': '=SUM({0}{1}:{0}{2})'}


class OfflineWriteReport(object):

    def __init__(self):
        pass

    def find_report_file(self, orderNo):
        for file in os.listdir('./report/'):
            if file.endswith('xlsx') and file.find(orderNo) >= 0:
                return './report/{0}'.format(file)
        return None

    def _init_srckey_dest_col_index_dict(self, destHeads, srcDestKeyDict):
        srcKeyDestColIndexMap = {}
        for srcKey, destKey in srcDestKeyDict.items():
            for index, destHead in enumerate(destHeads):
                if destHead and destHead.find(destKey) >= 0:
                    srcKeyDestColIndexMap.update({srcKey: index})
                    break
        return srcKeyDestColIndexMap

    def read_input_data(self, inputFilePath):
        f = open(inputFilePath)
        datas = []
        for row in csv.DictReader(f):
            if str(row[u'录入方式'.encode('gbk')]) == '1':
                datas.append(row)
        f.close()
        return datas

    def __get_heads(self, sheet, headIndex=0):
        rows = list(sheet.rows)
        heads = rows[headIndex]
        return map(lambda cell: cell.value, heads)

    def __init_formula_dict(self, sheet, formulaDict, headIndex=6):
        heads = self.__get_heads(sheet, headIndex)

        formulaItems = []
        for key, value in formulaDict.items():
            value = value.replace('+', '#').replace('-', '#').replace('*', '#').replace('/', '#')
            for colName in value.split('#'):
                formulaItems.append(colName)
            formulaItems.append(key)
        formulaItems = list(set(formulaItems))

        srcKeyDestColIndexMap = {}
        for formulaItem in formulaItems:
            for index, destHead in enumerate(heads):
                if destHead and destHead.find(formulaItem) >= 0:
                    srcKeyDestColIndexMap.update({formulaItem: index})
                    break

        transformedFormulaDict = {}
        for fKey, fValue in formulaDict.items():
            for key, value in srcKeyDestColIndexMap.items():
                fValue = fValue.replace(key, chr(value + 65))
            transformedFormulaDict.update({srcKeyDestColIndexMap.get(fKey): str(fValue)})
        return transformedFormulaDict

    def __init_col_index(self, sheet, srcDestKeyDict, headIndex=0):
        rows = list(sheet.rows)
        heads = rows[headIndex]
        heads = map(lambda cell: cell.value, heads)
        return self._init_srckey_dest_col_index_dict(heads, srcDestKeyDict)

    def __delete_null_rows(self, sheet, wb, outFilePath, dataBeginIndex=2):
        rows = list(sheet.rows)
        for rowIndex, row in enumerate(rows):
            if rowIndex < dataBeginIndex:
                continue
            if row[0].value is None:
                beginDelRowIndex = rowIndex
                sheet.delete_rows(beginDelRowIndex + 1, len(rows)-rowIndex)
                break
        wb.save(outFilePath)

    def _divide_input_datas(self, inputDatas):
        orderNoInputDatasDict = {}
        for inputData in inputDatas:
            orderNo = inputData.get(u'订单号'.encode('gbk'))
            orderNoInputDatasDict.setdefault(orderNo, []).append(inputData)
        return orderNoInputDatasDict

    def write_report(self, inputFilePath):
        inputDatas = self.read_input_data(inputFilePath)
        orderNoInputDatasDict = self._divide_input_datas(inputDatas)
        for orderNo, inputDatas in orderNoInputDatasDict.items():
            orderFilePath = self.find_report_file(orderNo)
            if orderFilePath is None:
                print 'report file for {0} not found!'.format(orderNo)
                continue

            wb = load_workbook(orderFilePath)
            sheet = wb.get_sheet_by_name(u'分选明细')

            self.__delete_null_rows(sheet, wb, orderFilePath)
            srcKeyDestColIndexMap = self.__init_col_index(sheet, SrcDestKeyDict)

            formulaDict = self.__init_formula_dict(sheet, FormulaDict, 0)

            dates = []
            for inputData in inputDatas:
                maxRow = sheet.max_row
                sheet.insert_rows(maxRow + 1)
                for srcKey, destColIndex in srcKeyDestColIndexMap.items():
                    dataType = self._get_data_type(srcKey)
                    sheet.cell(row=maxRow + 1, column=destColIndex + 1).value = self.__encode_data(inputData.get(srcKey.encode('gbk')), dataType)
                    sheet.cell(row=maxRow + 1, column=1+srcKeyDestColIndexMap.get(u'不良数量')).data_type = 'n'
                    self._set_row_boder(sheet, maxRow + 1)
                for col, formulaStr in formulaDict.items():
                    sheet.cell(row=maxRow + 1, column=col + 1).value = self._generate_formula(maxRow + 1, formulaStr)
                    self._set_row_boder(sheet, maxRow + 1)
                self._write_overall_cell(sheet, OverallFormulaDict, 2, 3, maxRow + 1, 0)
                dates.append(self.__encode_data(inputData.get(u'检验日期'.encode('gbk')), 'dateStr'))
            self._merge_cell(sheet)
            wb.save(orderFilePath)
            self._rename_date_in_filename(orderFilePath, max(dates))

    def _get_all_merge_cell(self, sheet):
        mergeCells = []
        for range in sheet.merged_cell_ranges:
            mergeCells.append((range.min_row, range.min_col, range.max_row, range.max_col))
        return mergeCells

    def _merge_cell(self, sheet):
        heads = self.__get_heads(sheet, 0)
        mergeColIndexs = []
        for mergeCol in MergeCols:
            for index, destHead in enumerate(heads):
                if destHead and destHead.find(mergeCol) >= 0:
                    mergeColIndexs.append(index + 1)
                    break
        col = mergeColIndexs[0]

        mergedCell = self._get_all_merge_cell(sheet)

        startRow = -1
        mergedRows = []
        for i in range(3, sheet.max_row + 1):
            if sheet.cell(i, col).value is not None and sheet.cell(i, col).value != '':
                if startRow > 0 and i - startRow > 1:
                    mergedRows.append([startRow, i - 1])
                startRow = i
        if sheet.max_row > startRow:
            mergedRows.append([startRow, sheet.max_row])

        for mergedRow in mergedRows:
            minRow = mergedRow[0]
            maxRow = mergedRow[1]
            for col in mergeColIndexs:
                if (minRow, col, maxRow, col) not in mergedCell:
                    sheet.merge_cells(None, minRow, col, maxRow, col)
                    sheet.cell(minRow, col).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')


    def _write_overall_cell(self, sheet, overallFormulaDict, cellRow, startRow, endRow, headIndex=6):
        heads = self.__get_heads(sheet, headIndex)
        for headName, formula in overallFormulaDict.items():
            for index, destHead in enumerate(heads):
                if destHead and destHead.find(headName) >= 0:
                    col = index + 1
                    sheet.cell(cellRow, col).value = formula.format(chr(index + 65), startRow, endRow)
                    break

    def _set_row_boder(self, sheet, rowIndex):
        for col in range(sheet.max_column):
            cell = sheet.cell(row=rowIndex, column=col + 1)
            cell.border = Border(left=Side(style='thin', color='000000'),
                                 right=Side(style='thin', color='000000'),
                                 top=Side(style='thin', color='000000'),
                                 bottom=Side(style='thin', color='000000'))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _generate_formula(self, row, formulaStr):
        for i in range(65, 65 + 26):
            formulaStr = formulaStr.replace(chr(i), '{0}{1}'.format(chr(i), row))
        return '=' + formulaStr

    def _get_data_type(self, srcKey):
        dataType = None
        if srcKey in [u'检验总数', u'不良数量']:
            dataType = int
        if srcKey in [u'任务工时', u'效率']:
            dataType = float
        if srcKey in [u'检验日期']:
            dataType = 'date'
        if srcKey in [u'批次号']:
            dataType = 'batchNo'
        return dataType

    def format_batch_no(self, batchNo):
        if re.search(ur'^[0-9]+$', batchNo):
            return str(batchNo.zfill(10))
        return batchNo.decode('gbk').encode('utf-8')

    def __encode_data(self, data, dataType=None):
        if dataType == 'date':
            m = re.search(u'(\d+).(\d+).(\d+)', data.decode('gbk'))
            if m:
                return datetime.strptime('{0}-{1}-{2}'.format(m.group(1), m.group(2), m.group(3)), '%Y-%m-%d').date()
        if dataType == 'dateStr':
            m = re.search(u'(\d+).(\d+).(\d+)', data.decode('gbk'))
            if m:
                return '{0}-{1}-{2}'.format(m.group(1), m.group(2), m.group(3))
        elif dataType == 'batchNo':
            return self.format_batch_no(data)
        elif dataType is not None:
            if data:
                return dataType(data)
        try:
            data = data.decode('gbk')
            return data.encode('utf-8')
        except:
            return data

    def _rename_date_in_filename(self, fileName, maxDate):
        modifiedFileName = re.sub('\d{4}-\d{2}-\d{2}', maxDate, fileName)
        if maxDate != modifiedFileName:
            os.rename(fileName, modifiedFileName)
        else:
            print 'modify filename failed, check whether the date in file name is latest'


if __name__ == '__main__':
    writeReport = OfflineWriteReport()
    writeReport.write_report('output.csv')
