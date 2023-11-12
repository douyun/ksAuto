# coding=utf-8
import csv
import os
import re
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Border, Side

SrcDestKeyDict = {u'检验日期': u'日期',
                   u'零件名称': u'零件名称',
                   u'零件号': u'零件号',
                   u'批次号': u'批次号',
                   u'供应商批次': u'供应商批次',
                   u'检验总数': u'检测数量',
                   u'任务人员': u'检验员',
                   u'任务时间': u'工作时间',
                   u'效率': u'效率',
                  u'任务工时': u'工时'}

CostReportSrcDestKeyDict = {u'服务明细': u'缺陷项目',
                            u'服务地点': u'工作地点'}


class OfflineWriteReport(object):

    def __init__(self):
        pass

    def find_report_file(self, orderNo):
        for file in os.listdir('./report/'):
            if file.endswith('xlsx') and file.find(orderNo) >= 0:
                return './report/{0}'.format(file)
        return None

    def _init_srckey_dest_col_index_dict(self, destHeads, srcDestKeyDict):
        srcKeyDestColIndexMap = {}
        for srcKey, destKey in srcDestKeyDict.items():
            for index, destHead in enumerate(destHeads):
                if destHead and destHead.find(destKey) >= 0:
                    srcKeyDestColIndexMap.update({srcKey: index})
                    break
        return srcKeyDestColIndexMap

    def read_input_data(self, inputFilePath):
        f = open(inputFilePath)
        datas = []
        for row in csv.DictReader(f):
            if str(row[u'录入方式'.encode('gbk')]) == '1':
                datas.append(row)
        f.close()
        return datas

    def __init_col_index(self, sheet, srcDestKeyDict, headIndex=6):
        rows = list(sheet.rows)
        heads = rows[headIndex]
        heads = map(lambda cell: cell.value, heads)
        print heads
        return self._init_srckey_dest_col_index_dict(heads, srcDestKeyDict)

    def __delete_null_rows(self, sheet, wb, outFilePath, dataBeginIndex=8):
        rows = list(sheet.rows)
        for rowIndex, row in enumerate(rows):
            if rowIndex < dataBeginIndex:
                continue
            if row[0].value is None:
                beginDelRowIndex = rowIndex
                sheet.delete_rows(beginDelRowIndex + 1, len(rows)-rowIndex)
                break
        wb.save(outFilePath)

    def _divide_input_datas(self, inputDatas):
        orderNoInputDatasDict = {}
        for inputData in inputDatas:
            orderNo = inputData.get(u'订单号'.encode('gbk'))
            orderNoInputDatasDict.setdefault(orderNo, []).append(inputData)
        return orderNoInputDatasDict

    def write_report(self, inputFilePath):
        inputDatas = self.read_input_data(inputFilePath)
        orderNoInputDatasDict = self._divide_input_datas(inputDatas)
        for orderNo, inputDatas in orderNoInputDatasDict.items():
            orderFilePath = self.find_report_file(orderNo)
            if orderFilePath is None:
                print 'report file for {0} not found!'.format(orderNo)
                continue

            wb = load_workbook(orderFilePath)
            quantitySheet = wb.get_sheet_by_name(u'质量报表（report）')
            costSheet = wb.get_sheet_by_name(u'费用报表（cost list）')
            self.__delete_null_rows(quantitySheet, wb, orderFilePath)
            self.__delete_null_rows(costSheet, wb, orderFilePath, 6)
            srcKeyDestColIndexMap = self.__init_col_index(quantitySheet, SrcDestKeyDict, 6)
            costReportSrcKeyDestColIndexMap = self.__init_col_index(costSheet, CostReportSrcDestKeyDict, 4)
            for inputData in inputDatas:
                quantityMaxRow = quantitySheet.max_row
                costMaxRow = costSheet.max_row
                quantitySheet.insert_rows(quantityMaxRow + 1)
                costSheet.insert_rows(costMaxRow + 1)
                for srcKey, destColIndex in srcKeyDestColIndexMap.items():
                    dataType = self._get_data_type(srcKey)
                    quantitySheet.cell(row=quantityMaxRow + 1, column=destColIndex + 1).value = self.__encode_data(inputData.get(srcKey.encode('gbk')), dataType)
                    self._set_row_boder(quantitySheet, quantityMaxRow + 1)
                for srcKey, destColIndex in costReportSrcKeyDestColIndexMap.items():
                    costSheet.cell(row=costMaxRow + 1, column=destColIndex + 1).value = self.__encode_data(inputData.get(srcKey.encode('gbk')))
            wb.save(orderFilePath)

    def _set_row_boder(self, sheet, rowIndex):
        for col in range(sheet.max_column):
            sheet.cell(row=rowIndex, column=col+1).border = Border(left=Side(style='thin', color='000000'),
                                                                   right=Side(style='thin', color='000000'),
                                                                   top=Side(style='thin', color='000000'),
                                                                   bottom=Side(style='thin', color='000000'))
    def _get_data_type(self, srcKey):
        dataType = None
        if srcKey in [u'检验总数']:
            dataType = int
        if srcKey in [u'任务工时']:
            dataType = float
        if srcKey in [u'检验日期']:
            dataType = 'date'
        if srcKey in [u'批次号']:
            dataType = 'batchNo'
        return dataType

    def format_batch_no(self, batchNo):
        if re.search(ur'^[0-9]+$', batchNo):
            return str(batchNo.zfill(10))
        return batchNo.decode('gbk').encode('utf-8')

    def __encode_data(self, data, dataType=None):
        if dataType == 'date':
            m = re.search(u'(\d+).(\d+).(\d+)', data.decode('gbk'))
            if m:
                return datetime.strptime('{0}-{1}-{2}'.format(m.group(1), m.group(2), m.group(3)), '%Y-%m-%d').date()
        elif dataType == 'batchNo':
            return self.format_batch_no(data)
        elif dataType is not None:
            if data:
                return dataType(data)
        try:
            data = data.decode('gbk')
            return data.encode('utf-8')
        except:
            return data


if __name__ == '__main__':
    writeReport = OfflineWriteReport()
    writeReport.write_report('output.csv')
